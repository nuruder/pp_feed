"""
Export data to Excel.

Test independently:
    python -m export.excel [--output path.xlsx]
"""

import asyncio
import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, desc
from sqlalchemy.orm import selectinload

from db.database import AsyncSessionLocal, init_db
from db.models import Product, PriceSnapshot, Category

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


async def export_to_excel(output_path: str | Path | None = None) -> Path:
    """Export all products with latest prices to Excel."""
    await init_db()

    if output_path is None:
        output_path = Path("data") / f"padelpoint_export_{datetime.now():%Y%m%d_%H%M}.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Products sheet ---
    ws = wb.active
    ws.title = "Products"

    headers = [
        "ID", "External ID", "Name", "Brand", "Product Type", "Category",
        "Price (Regular)", "Price (Original)", "Price (Special)", "Price (Wholesale)",
        "Price (No Tax)", "Stock Qty", "In Stock", "Sizes",
        "URL", "Last Updated",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Product)
            .options(
                selectinload(Product.brand),
                selectinload(Product.product_type),
                selectinload(Product.categories),
                selectinload(Product.sizes),
            )
            .order_by(Product.name)
        )
        products = result.scalars().unique().all()

        row = 2
        for product in products:
            # Get latest snapshot
            snap_result = await session.execute(
                select(PriceSnapshot)
                .where(PriceSnapshot.product_id == product.id)
                .order_by(desc(PriceSnapshot.timestamp))
                .limit(1)
            )
            latest = snap_result.scalar_one_or_none()

            sizes_str = ", ".join(
                f"{s.size_label}({'v' if s.in_stock else 'x'})"
                for s in product.sizes
            )

            ws.cell(row=row, column=1, value=product.id)
            ws.cell(row=row, column=2, value=product.external_id)
            ws.cell(row=row, column=3, value=product.name)
            ws.cell(row=row, column=4, value=product.brand.name if product.brand else "")
            ws.cell(row=row, column=5, value=product.product_type.name if product.product_type else "")
            ws.cell(row=row, column=6, value=", ".join(c.name for c in product.categories) if product.categories else "")
            ws.cell(row=row, column=7, value=latest.price_regular if latest else None)
            ws.cell(row=row, column=8, value=latest.price_original if latest else None)
            ws.cell(row=row, column=9, value=latest.price_special if latest else None)
            ws.cell(row=row, column=10, value=latest.price_wholesale if latest else None)
            ws.cell(row=row, column=11, value=latest.price_without_tax if latest else None)
            ws.cell(row=row, column=12, value=latest.stock_quantity if latest else 0)
            ws.cell(row=row, column=13, value="Yes" if (latest and latest.in_stock) else "No")
            ws.cell(row=row, column=14, value=sizes_str)
            ws.cell(row=row, column=15, value=product.url)
            ws.cell(row=row, column=16, value=latest.timestamp.isoformat() if latest else "")
            row += 1

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # --- Categories sheet ---
    ws2 = wb.create_sheet("Categories")
    cat_headers = ["ID", "Name", "Parent", "Level", "URL"]
    for col, header in enumerate(cat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    async with AsyncSessionLocal() as session:
        result = await session.execute(
            select(Category)
            .options(selectinload(Category.parent))
            .order_by(Category.level, Category.name)
        )
        categories = result.scalars().unique().all()

        for i, cat in enumerate(categories, 2):
            ws2.cell(row=i, column=1, value=cat.id)
            ws2.cell(row=i, column=2, value=cat.name)
            ws2.cell(row=i, column=3, value=cat.parent.name if cat.parent else "")
            ws2.cell(row=i, column=4, value=cat.level)
            ws2.cell(row=i, column=5, value=cat.url)

    wb.save(output_path)
    logger.info("Exported to %s (%d products)", output_path, row - 2)
    return output_path


async def run():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    output = None
    for i, arg in enumerate(sys.argv):
        if arg == "--output" and i + 1 < len(sys.argv):
            output = sys.argv[i + 1]
    path = await export_to_excel(output)
    print(f"Export saved to: {path}")


if __name__ == "__main__":
    asyncio.run(run())
